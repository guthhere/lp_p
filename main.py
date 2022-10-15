import attr
import requests
import ast
from bs4 import BeautifulSoup
import browsercookie
import js2py
import openpyxl
from openpyxl import Workbook
import os
import pandas
from datetime import datetime
import random
import time

art = {"response":[{"productId":15162299,"rating":4.3,"reviews_number":20},{"productId":17932103,"rating":3.9,"reviews_number":17},{"productId":12844761,"rating":3,"reviews_number":5},{"productId":17932091,"rating":4,"reviews_number":3},{"productId":18623548,"rating":3.9,"reviews_number":22},{"productId":12845895,"rating":3.1,"reviews_number":10},{"productId":18623521,"rating":4.8,"reviews_number":8},{"productId":15090301,"rating":4.4,"reviews_number":12},{"productId":18623505,"rating":3.4,"reviews_number":5},{"productId":18623492,"rating":3.2,"reviews_number":7},{"productId":18623530,"rating":3.2,"reviews_number":16},{"productId":17932058,"rating":4.2,"reviews_number":5},{"productId":15090310,"rating":4.4,"reviews_number":6},{"productId":17932031,"rating":5,"reviews_number":2},{"productId":12719418,"rating":4,"reviews_number":4},{"productId":18623556,"rating":4.5,"reviews_number":2},{"productId":17932040,"rating":4,"reviews_number":1},{"productId":17932111,"rating":3.8,"reviews_number":4},{"productId":18623581,"rating":3,"reviews_number":8},{"productId":12846169,"rating":3.2,"reviews_number":6},{"productId":12846150,"rating":3.6,"reviews_number":5},{"productId":12845983,"rating":4,"reviews_number":5},{"productId":12845887,"rating":5,"reviews_number":1},{"productId":18623601,"rating":5,"reviews_number":3},{"productId":17932066,"rating":4,"reviews_number":3},{"productId":18648745,"rating":4.6,"reviews_number":7},{"productId":18623847,"rating":0,"reviews_number":0},{"productId":18623513,"rating":5,"reviews_number":3},{"productId":18623572,"rating":3.8,"reviews_number":12},{"productId":12845852,"rating":0,"reviews_number":0}]}

header = {"Accept": "application/json, text/plain, */*",
"Accept-Encoding": "gzip, deflate, br",
"Accept-Language": "en-US,en;q=0.6",
"Connection": "keep-alive",
"Content-Length": "114",
"Content-Type": "application/json;charset=UTF-8",
"Cookie": "disp_react_aa=1; ggr-widget-test=0; disp_delivery_ab=A; cookie_accepted=true; sid=NU6TtCw4r0dlFCiwxeTYM0ePunST4-RV.PZC4qshANtzrg4DIQzHRWlR6lMMAf5AJd9lLb4jZy4U; iap.uid=47d69bdbc39247558942fba9f1c3054b; fromRegion=34; lastConfirmedRegionID=34; _regionID=506",
"Host": "api.leroymerlin.ru",
"Origin": "https://spb.leroymerlin.ru",
"Referer": "https://spb.leroymerlin.ru/",
"Sec-Fetch-Dest": "empty",
"Sec-Fetch-Mode": "cors",
"Sec-Fetch-Site": "same-site",
"Sec-GPC": "1",
"User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/106.0.0.0 Safari/537.36",
"x-api-key": "nkGKLkscp80GVAQVY8YvajPjzaFTmIS8",
"x-request-id": "229e608adf27002566034921f4d4e680",}

cj = browsercookie.chrome()


def get_data_petrovich():
    data = {}
    ses = requests.Session()
    cj = browsercookie.chrome()
    r = ses.get('https://petrovich.ru/catalog/15635/', cookies=cj)
    soup = BeautifulSoup(r.text, 'lxml')
    print(r.status_code)

    ids = []
    ids += [product.find('p', attrs={'data-test': 'product-code'}).getText()
            for product in soup.find_all('div', class_='pt-flex pt-flex-col pt-items-center')]
    r = ses.get('https://petrovich.ru/catalog/15635/?p=1', cookies=cj)
    soup = BeautifulSoup(r.text, 'lxml')
    ids += [product.find('p', attrs={'data-test': 'product-code'}).getText()
            for product in soup.find_all('div', class_='pt-flex pt-flex-col pt-items-center')]
    print(ids)
    for script in soup.find_all('script'):
        if script.get('src') != None:
            if 'init.js' in script.get('src'):
                init_url = script.get('src')

    # print(ses.cookies)
    #ses.headers = headers
    r = ses.get(init_url)
    # print(ses.headers)
    for id in ids:
        r = ses.get(
            f'https://api.petrovich.ru/catalog/v2.3/products/{id}?city_code=spb&client_id=pet_site')
        # print(ses.cookies)
        # print(ses.headers)
        # print(r.json()['data']['product'].keys())
        data[int(id)] = r.json()['data']['product']['remains']['supply_ways'][0]['total']

    return data


def get_data_leroy():
    ses = requests.Session()
    r = ses.get("https://spb.leroymerlin.ru",cookies=cj).text
    print(r)
    if '/__qrator/qauth_utm_v2.js' in r:
        r = ses.get("https://leroymerlin.ru/__qrator/qauth_utm_v2.js",cookies=cj).text
    #js2py.eval_js(r)
    id = 15162299
    ids =[id['productId'] for id in art['response']]
    #print(r)
    print(ids)
    data={}
    for id in ids:
        url = "https://api.leroymerlin.ru/experience/LeroymerlinWebsite/v1/navigation-pdp-api//get-stocks"
        

        payload = {"currencyKey": "RUB",
                "preferedStores": [],
                "productId": str(id),
                "regionCode": "spb",
                "source": "Step",
                "unit": "шт."}
        r = ses.post(url, json = payload, headers = header)
      
        #print(r.text.json())
        sum = 0
        try:
            for stock in r.json()['stocks']:
                sum+=stock['stockValue'] 
            print(sum)
            data[id] = sum
        except:
            print(r.status_code)
            print(r.headers)
            print(r.text)
    return data
def create_file(petr_data,leroy_data):
    now = datetime.now()
 

    date = now.strftime("%d/%m/%Y %H:%M:%S")
    if os.path.isfile('hello_world.xlsx') == False:
        workbook = Workbook()
        sheet = workbook.active
        i=0
        sheet.cell(row=1, column=1, value='ТИП')
        sheet.cell(row=1, column=2, value='КОД')
        sheet.cell(row=1, column=3, value=date)
        for id in petr_data:
            sheet.cell(row=i+2, column=1, value='VEKA VHS Halo')
            sheet.cell(row=i+2, column=2, value=id)
            sheet.cell(row=i+2, column=3, value=petr_data[id])
            i+=1
        i=0
        for id in leroy_data:
            sheet.cell(row=i+len(petr_data)+1, column=1, value='Deceuninck')
            sheet.cell(row=i+len(petr_data)+1, column=2, value=id)
            sheet.cell(row=i+len(petr_data)+1, column=3, value=leroy_data[id])
            i+=1
        workbook.save(filename="hello_world.xlsx")  

    else:
        workbook = openpyxl.load_workbook('hello_world.xlsx')
        data = petr_data | leroy_data
        sheet = workbook.active
        ids = len(sheet['A'])
        col_min = len(sheet['1'])
        sheet.cell(row=1, column=col_min+1, value='-->')
        sheet.cell(row=1, column=col_min+2, value=date)
        for i in range(ids):
            value =  sheet.cell(row = i+2, column=col_min).value
            id =  sheet.cell(row = i+2, column=2).value
            if id in data:
                if value <= data[id]:
                    sheet.cell(row=i+2, column=col_min+1, value=0)
                    sheet.cell(row=i+2, column=col_min+2, value=data[id])
                else:
                    sheet.cell(row=i+2, column=col_min+1, value=value- data[id])
                    sheet.cell(row=i+2, column=col_min+2, value=data[id])
            else:
                sheet.cell(row=i+1, column=col_min+1, value=0)
                sheet.cell(row=i+1, column=col_min+2, value=value)
            
        #rows= ((11,22), (333,43))
        #sheet.cell(row=1, column=10, value='date')
        #or row in rows:
         #   sheet.append(row)

        workbook.save(filename="hello_world.xlsx")  
if __name__ == '__main__':
    petr_data=[681102,167176,618692,167165,681111,167166,681113,681104,167167,681112,681103,681116,681114,618688,618686,681122,681118,681121,618687,681115,681117,681119,681125,618691,681120,681106,681105]
    leroy_data=[15162299, 17932103, 12844761, 17932091, 18623548, 12845895, 18623521, 15090301, 18623505, 18623492, 18623530, 17932058, 15090310, 17932031, 12719418, 18623556, 17932040, 17932111, 18623581, 12846169, 12846150, 12845983, 12845887, 18623601, 17932066, 18648745, 18623847, 18623513, 18623572, 12845852]
    petr_data = get_data_petrovich()
    leroy_data = get_data_leroy()
    #print(petr_data)
    #print(leroy_data)
    create_file(petr_data,leroy_data)
    
